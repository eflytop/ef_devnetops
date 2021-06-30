#!/usr/bin/env python3 
# -*- coding=utf-8 -*-
# by EFLYTOP
# www.eflytop.com

from netmiko import ConnectHandler,NetmikoTimeoutException,NetmikoAuthenticationException 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from pprint import pprint
from concurrent.futures import ThreadPoolExecutor
import getpass
import sys
import time

username = input("请输入设备用户名：")
password = getpass.getpass(prompt='请输入密码:')
#enpass = getpass.getpass(prompt='请输入enable密码:')

succeed_ip_list = []
failed_login_ip_list = []
hostname_list = []
sn_list = []
uptime_list = []
model_list = []
os_version_list = []
image_list = []

wb = Workbook()
ws = wb.active
ws.title = 'Inventory'
ws['A1'] = 'Hostname'
ws['B1'] = 'IP Address'
ws['C1'] = 'Serial Number'
ws['D1'] = 'Uptime'
ws['E1'] = 'Model'
ws['F1'] = 'OS Version'
ws['G1'] = 'Image'
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws['A1'].fill=yellowFill
ws['B1'].fill=yellowFill
ws['C1'].fill=yellowFill
ws['D1'].fill=yellowFill
ws['E1'].fill=yellowFill
ws['F1'].fill=yellowFill
ws['G1'].fill=yellowFill

#print输出到txt文件
class Logger(object):
	def __init__(self, filename="Default.log"):
		self.terminal = sys.stdout
		self.log = open(filename, "a")
	def write(self, message):
		self.terminal.write(message)
		self.log.write(message)
	def flush(self):
		pass
sys.stdout = Logger('log.txt')

start_time = time.time()

def get_ip_address():
    with open('device.txt') as f:
        addresses = f.read().splitlines()
    return addresses

def retrieve_data(ipaddr):
    connection_info = {
	    	'device_type': 'cisco_ios',
	    	'ip': ipaddr,
	    	'username': username,
	   		'password': password,
	    	#'port': 22,
	   	 	#'secret': enpass,
    }
    try:
        with ConnectHandler(**connection_info) as conn:
            print('INFO: ',time.strftime('%x %X'),'已经登录设备：', ipaddr)
            output = conn.send_command('show version', use_textfsm=True)
            succeed_ip_list.append(ipaddr)
            hostname = conn.find_prompt().replace('#','')
            hostname_list.append(hostname)
            sn = output[0]['serial'][0]
            sn_list.append(sn)
            uptime = output[0]['uptime']
            uptime_list.append(uptime)
            model = output[0]['hardware'][0]
            model_list.append(model)
            os_version = output[0]['version']
            os_version_list.append(os_version)
            image = output[0]['running_image']
            image_list.append(image)

    except NetmikoAuthenticationException : #认证失败报错记录
    	print('INFO: ',time.strftime('%x %X'),ipaddr,'[Error 1] Authentication failed.')
    	failed_login_ip_list.append(ipaddr)

    except NetmikoTimeoutException : #登录超时报错记录
    	print('INFO: ',time.strftime('%x %X'),ipaddr,'[Error 2] Connection timed out.')
    	failed_login_ip_list.append(ipaddr)

    except Exception as e:
    	print('INFO: ',time.strftime('%x %X'),ipaddr,e)
    	failed_login_ip_list.append(ipaddr)

with ThreadPoolExecutor(max_workers=5) as exe:
    ip_addresses = get_ip_address()
    results = exe.map(retrieve_data, ip_addresses)


row_top_number = len(succeed_ip_list) + 2
for hostname, ip, sn, uptime, model, os_version, image, row in zip(hostname_list, succeed_ip_list, sn_list, uptime_list, model_list, os_version_list, image_list, range(2, row_top_number)):
    ws.cell(row=row, column=1, value=hostname)
    ws.cell(row=row, column=2, value=ip)
    ws.cell(row=row, column=3, value=sn)
    ws.cell(row=row, column=4, value=uptime)
    ws.cell(row=row, column=5, value=model)
    ws.cell(row=row, column=6, value=os_version)
    ws.cell(row=row, column=7, value=image)

dims = {}
for row in ws.rows:
    for cell in row:
        cell.border=thin_border
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column, 0), len(str(cell.value))))

for col, value in dims.items():
    ws.column_dimensions[col].width = value + 5

wb.save('inventory.xlsx')

print('下列主机登录失败，请手动检查：')
for i in failed_login_ip_list:
	print(i)
print(succeed_ip_list)
print(hostname_list)
print(model_list)
print('本次执行时长：%.2f'%(time.time()-start_time) + '秒')
print("命令执行完毕，结果请查看log.txt文件")